from openpyxl import load_workbook
import os

def formata_planilha(nome_arquivo='contatos'):
    nome_arquivo.lower()
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    CAMINHO_PLANILHAS = os.path.join(BASE_DIR, 'planilha', f'{nome_arquivo}.xlsx')
    CAMINHO_PLANILHAS = os.path.normpath(CAMINHO_PLANILHAS)

    


    PATH = CAMINHO_PLANILHAS
    # Abre a planilha
    wb = load_workbook(PATH)
    ws = wb.active

    # Pega só cabeçalhos válidos (não nulos)
    headers = [cell.value for cell in ws[1] if cell.value]

    contatos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = row[:len(headers)]

        # Ignora linhas totalmente vazias
        if all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row_data):
            continue

        # Monta o dicionário apenas para colunas com título
        contato = dict(zip(headers, row_data))
        contato = {k: (v.strip() if isinstance(v, str) else v) for k, v in contato.items()}

        contatos.append(contato)

    return contatos



WHATSAPP_PATH = {"url":"https://web.whatsapp.com/",   "chrome_path":"C:/Program Files/Google/Chrome/Application/chrome.exe %s"}

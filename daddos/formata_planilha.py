from openpyxl import load_workbook

def formata_planilha():
    caminho = 'C:\\Users\\supor\\Documents\\boot-mfprint\\daddos\\planilha\\Contatos.xlsx'
    wb = load_workbook(caminho)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    contatos = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # começa da linha 2
        contato = dict(zip(headers, row))
        contatos.append(contato)

    return contatos
   

print(formata_planilha())